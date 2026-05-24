"""Build fixtures/iter5_stress.xlsx — the cohort that historically tripped
blank-crop verify-rejection. Iter-5's Pass-3 retry + variance pre-gate should
recover a chunk of these.

Layout matches demo_workbook.xlsx exactly (sheet `图形商标tro`, columns A-K).
The pipeline only needs columns B (appno), D (logo_url), K (evidence_url, "\n"-joined).
"""
import json
import sqlite3
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
DB = ROOT / ".data" / "linebase.db"
SRC = ROOT / "fixtures" / "demo_workbook.xlsx"
DST = ROOT / "fixtures" / "iter5_stress.xlsx"

# Iter-5 stress targets — appnos that historically had blank-crop verify rejects
# plus two clean regression baselines.
STRESS = [
    "75537343", "75537827", "87135634", "77354840", "74677565",
    "97764930", "97975988", "85094272", "79048263",
]
BASELINE = ["78402423", "74677567"]
WANT = STRESS + BASELINE


def fetch_row_from_db(appno: str) -> tuple[str, list[str]]:
    conn = sqlite3.connect(DB)
    r = conn.execute(
        "SELECT logo_url, evidence_urls_json FROM job_row WHERE appno=? LIMIT 1",
        (appno,),
    ).fetchone()
    if not r:
        raise SystemExit(f"appno {appno} not in DB")
    return r[0], json.loads(r[1] or "[]")


def main() -> None:
    src_wb = openpyxl.load_workbook(SRC)
    src_ws = src_wb["图形商标tro"]

    # Index demo_workbook by appno
    by_appno: dict[str, tuple] = {}
    header_row = None
    for i, row in enumerate(src_ws.iter_rows(values_only=True), start=1):
        if i == 1:
            header_row = row
            continue
        appno = str(row[1]) if row[1] is not None else None
        if appno:
            by_appno[appno] = row

    dst_wb = openpyxl.Workbook()
    dst_ws = dst_wb.active
    dst_ws.title = "图形商标tro"
    dst_ws.append(list(header_row))

    for appno in WANT:
        if appno in by_appno:
            dst_ws.append(list(by_appno[appno]))
            print(f"  [{appno}] copied from demo_workbook")
        else:
            logo_url, evs = fetch_row_from_db(appno)
            # build a minimal row matching column layout
            new_row = [None] * len(header_row)
            new_row[1] = appno
            new_row[3] = logo_url
            new_row[10] = "\n".join(evs)
            dst_ws.append(new_row)
            print(f"  [{appno}] built from DB ({len(evs)} evidence URLs)")

    dst_wb.save(DST)
    print(f"\nWrote {DST}")
    print(f"Rows: {len(WANT)}  (stress={len(STRESS)} baseline={len(BASELINE)})")


if __name__ == "__main__":
    main()
