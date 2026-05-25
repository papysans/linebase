"""Build a fixtures xlsx mapping truth-set pairs to http URLs served by the
static server on http://127.0.0.1:8001/. Each row = one TM with logo URL + the
TM's first 3 evidence URLs (so we cap per-row LLM calls)."""
import json
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
SRC_DEMO = ROOT / "fixtures" / "demo_workbook.xlsx"
DST = ROOT / "fixtures" / "truthset_e2e.xlsx"
STATIC_BASE = "http://127.0.0.1:8001/truth_set"
INDEX = json.loads((ROOT / "docs" / "truth_set" / "INDEX.json").read_text(encoding="utf-8"))

# Copy header from demo_workbook to keep column layout identical (D=logo, K=evidence)
demo_wb = openpyxl.load_workbook(SRC_DEMO)
demo_ws = demo_wb["图形商标tro"]
header = list(next(demo_ws.iter_rows(values_only=True)))

dst_wb = openpyxl.Workbook()
ws = dst_wb.active
ws.title = "图形商标tro"
ws.append(header)

for block in INDEX:
    tm = block["tm"]
    cat = block["cat"]
    logo_rel = block["logo"]  # e.g. docs/truth_set/2423810_30/logo.png
    logo_url = f"{STATIC_BASE}/{Path(logo_rel).relative_to('docs/truth_set').as_posix()}"

    # Pick all pairs that have a reliable truth bbox (score >= 0.85)
    ev_urls = []
    for p in block["pairs"]:
        if (p.get("score") or 0) < 0.85 or not p.get("truth_bbox"):
            continue
        ev_rel = p["evidence"]
        ev_url = f"{STATIC_BASE}/{Path(ev_rel).relative_to('docs/truth_set').as_posix()}"
        ev_urls.append(ev_url)

    row = [None] * len(header)
    row[0] = "是"
    row[1] = tm
    row[3] = logo_url
    row[4] = cat
    row[10] = ",".join(ev_urls)
    ws.append(row)
    print(f"  [{tm}] cat={cat} logo={logo_url[:60]}... ev={len(ev_urls)}")

dst_wb.save(DST)
print(f"\nWrote {DST}, {len(INDEX)} rows")
