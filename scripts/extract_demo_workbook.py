"""Extract a slimmed canonical demo xlsx from the 428 MB source workbook.

Reads only the `图形商标tro` sheet, preserves the header row, and walks
data rows from row 2 onwards keeping the first 10 rows where:
  - column B (申请号) is non-empty
  - column D (logo URL) starts with "http"
  - column K (使用证据) contains at least one comma-separated http URL

Writes a fresh workbook to fixtures/demo_workbook.xlsx (~50-200 KB), preserves
the 11-column header layout (A-K), and prints a one-shot summary.

Usage:
    .venv/Scripts/python.exe -u scripts/extract_demo_workbook.py
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
from openpyxl import Workbook

REPO = Path(__file__).resolve().parents[1]
SOURCE = REPO / "美专实物图排查-2026.2.6.xlsx"
SHEET_NAME = "图形商标tro"
OUT_PATH = REPO / "fixtures" / "demo_workbook.xlsx"
# NOTE: the PRD originally assumed the source had 2 header rows (row 1 + a Chinese
# sub-header on row 2). Inspection of the real file shows there is only ONE header
# row (row 1: 申请号 / 图形商标logo / 使用证据), and row 2 onwards is real data.
# The downstream pipeline (`iter_rows` in io_excel.py) also defaults to start_row=2,
# i.e. assumes a single header. Preserving 2 header rows would corrupt the format
# contract for downstream consumers, so we preserve exactly 1.
TARGET_ROWS = 10
HEADER_ROW_COUNT = 1
NUM_COLS = 11  # A..K


def _row_has_evidence_urls(value: object) -> int:
    """Return the count of comma-separated http URLs in a K-column value."""
    if not isinstance(value, str):
        return 0
    parts = [p.strip() for p in value.split(",")]
    return sum(1 for p in parts if p.lower().startswith("http"))


def main() -> int:
    if not SOURCE.exists():
        print(f"[FAIL] source workbook not found: {SOURCE}")
        return 2

    print(f"[info] opening {SOURCE.name} (read_only, data_only)...")
    wb = openpyxl.load_workbook(SOURCE, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        print(f"[FAIL] sheet not found: {SHEET_NAME}")
        print(f"       available: {wb.sheetnames}")
        return 2
    ws = wb[SHEET_NAME]

    # Stream rows; capture the two header rows and the first 10 qualifying data rows.
    headers: list[list[object]] = []
    picked: list[list[object]] = []
    evidence_counts: list[int] = []
    appnos: list[str] = []
    scanned = 0

    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        # Normalize to NUM_COLS columns (pad / truncate).
        cells = list(row[:NUM_COLS])
        if len(cells) < NUM_COLS:
            cells.extend([None] * (NUM_COLS - len(cells)))

        if idx <= HEADER_ROW_COUNT:
            headers.append(cells)
            continue

        scanned += 1
        appno = cells[1]  # column B
        logo_url = cells[3]  # column D
        evidence = cells[10]  # column K

        if appno is None or str(appno).strip() == "":
            continue
        if not (isinstance(logo_url, str) and logo_url.strip().lower().startswith("http")):
            continue
        ec = _row_has_evidence_urls(evidence)
        if ec < 1:
            continue

        picked.append(cells)
        evidence_counts.append(ec)
        appnos.append(str(appno).strip())

        if len(picked) >= TARGET_ROWS:
            break

    wb.close()

    if len(picked) < TARGET_ROWS:
        print(f"[FAIL] only found {len(picked)} qualifying rows after scanning {scanned}; need {TARGET_ROWS}")
        return 2

    # Write fresh workbook.
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    out_wb = Workbook()
    # Replace the default sheet with our target name.
    out_ws = out_wb.active
    out_ws.title = SHEET_NAME

    for cells in headers:
        out_ws.append(cells)
    for cells in picked:
        out_ws.append(cells)

    # Reasonable column widths: 申请号 ~ 12, URL columns ~ 60, others default 18.
    width_map = {
        "A": 18,  # row id / leading column
        "B": 12,  # 申请号
        "C": 18,
        "D": 60,  # logo URL
        "E": 18,
        "F": 18,
        "G": 18,
        "H": 18,
        "I": 18,
        "J": 18,
        "K": 60,  # evidence URLs
    }
    for col, width in width_map.items():
        out_ws.column_dimensions[col].width = width

    out_wb.save(OUT_PATH)
    out_wb.close()

    size_bytes = OUT_PATH.stat().st_size
    avg_ev = sum(evidence_counts) / len(evidence_counts)

    print()
    print(f"[OK] wrote {OUT_PATH}")
    print(f"     size: {size_bytes} bytes ({size_bytes / 1024:.1f} KB)")
    print(f"     scanned {scanned} data rows to find {len(picked)} qualifying")
    print(f"     sheet: {SHEET_NAME}")
    print(f"     layout: {HEADER_ROW_COUNT} header rows + {len(picked)} data rows = {HEADER_ROW_COUNT + len(picked)} rows total, {NUM_COLS} columns (A..K)")
    print(f"     average evidence URL count per row: {avg_ev:.2f}")
    print(f"     申请号 list ({len(appnos)}):")
    for i, a in enumerate(appnos, start=1):
        print(f"       {i:2d}. {a}  (evidence_count={evidence_counts[i - 1]})")

    if size_bytes > 200 * 1024:
        print(f"[warn] output is {size_bytes / 1024:.1f} KB, larger than the ~50-200 KB target")
    return 0


if __name__ == "__main__":
    sys.exit(main())
