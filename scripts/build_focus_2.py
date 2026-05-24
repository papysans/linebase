"""Tiny 2-row stress xlsx for the failing samples 87135634 + 97457879."""
import json
import sqlite3
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
DB = ROOT / ".data" / "linebase.db"
SRC = ROOT / "fixtures" / "demo_workbook.xlsx"
DST = ROOT / "fixtures" / "focus_2failures.xlsx"

TARGET = ["87135634", "97457879"]

src_wb = openpyxl.load_workbook(SRC)
src_ws = src_wb["图形商标tro"]

by_appno = {}
header_row = None
for i, row in enumerate(src_ws.iter_rows(values_only=True), start=1):
    if i == 1:
        header_row = row
        continue
    if row[1] is not None:
        by_appno[str(row[1])] = row

dst_wb = openpyxl.Workbook()
dst_ws = dst_wb.active
dst_ws.title = "图形商标tro"
dst_ws.append(list(header_row))

conn = sqlite3.connect(DB)
for appno in TARGET:
    if appno in by_appno:
        dst_ws.append(list(by_appno[appno]))
        print(f"  [{appno}] from demo_workbook")
    else:
        r = conn.execute("SELECT logo_url, evidence_urls_json FROM job_row WHERE appno=? LIMIT 1", (appno,)).fetchone()
        new_row = [None] * len(header_row)
        new_row[1] = appno
        new_row[3] = r[0]
        new_row[10] = "\n".join(json.loads(r[1] or "[]"))
        dst_ws.append(new_row)
        print(f"  [{appno}] from DB")

dst_wb.save(DST)
print(f"\nWrote {DST}")
