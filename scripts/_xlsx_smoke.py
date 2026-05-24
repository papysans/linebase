"""Smoke test: GET /api/jobs/{id}/xlsx returns a real xlsx (PK zip magic).

Run against a live uvicorn on :8765. Picks the most recent finished job from
/api/jobs and asserts:
  - HTTP 200
  - Content-Type is the xlsx MIME
  - Body starts with the zip magic `PK\\x03\\x04`
  - openpyxl can open it and the workbook has a `results` sheet

Also verifies the SPA catch-all does NOT swallow malformed /api/ paths
(regression for the "downloaded file is HTML" symptom): /api/jobs//xlsx with
an empty jobId must return HTTP 404, not 200 + index.html.

Usage:
    python scripts/_xlsx_smoke.py [base_url]
Exit code: 0 on pass, 1 on any failure.
"""
from __future__ import annotations

import io
import sys
import urllib.request

import openpyxl

BASE = sys.argv[1] if len(sys.argv) > 1 else "http://127.0.0.1:8765"


def _get(path: str, *, allow_404: bool = False) -> tuple[int, dict[str, str], bytes]:
    req = urllib.request.Request(f"{BASE}{path}")
    try:
        with urllib.request.urlopen(req, timeout=30) as r:
            return r.status, dict(r.headers), r.read()
    except urllib.error.HTTPError as e:
        if allow_404 and e.code == 404:
            return e.code, dict(e.headers or {}), b""
        raise


def main() -> int:
    import json

    status, _hdrs, body = _get("/api/jobs?limit=50")
    assert status == 200, f"GET /api/jobs failed: {status}"
    jobs = json.loads(body)
    finished = [j for j in jobs if j.get("status") == "finished" and j.get("done_rows", 0) > 0]
    if not finished:
        print("SKIP: no finished jobs with results to test against")
        return 0
    job_id = finished[0]["id"]
    print(f"Testing xlsx download for job {job_id} ...")

    status, hdrs, body = _get(f"/api/jobs/{job_id}/xlsx")
    assert status == 200, f"xlsx download returned {status}"
    ctype = hdrs.get("content-type") or hdrs.get("Content-Type") or ""
    assert "spreadsheetml.sheet" in ctype, f"wrong content-type: {ctype!r}"
    cdisp = hdrs.get("content-disposition") or hdrs.get("Content-Disposition") or ""
    assert "attachment" in cdisp and ".xlsx" in cdisp, f"wrong content-disposition: {cdisp!r}"
    assert body[:4] == b"PK\x03\x04", f"not a zip: first 8 bytes = {body[:8]!r}"
    wb = openpyxl.load_workbook(io.BytesIO(body))
    assert wb.sheetnames == ["results"], f"unexpected sheets: {wb.sheetnames}"
    print(f"OK: {len(body)} bytes, ctype={ctype}, sheets={wb.sheetnames}")

    # Regression: SPA catch-all must not swallow /api/ paths
    status, _hdrs, _body = _get("/api/jobs//xlsx", allow_404=True)
    assert status == 404, f"empty-jobId /api/jobs//xlsx should be 404, got {status}"
    print("OK: SPA catch-all rejects malformed /api/ paths")
    return 0


if __name__ == "__main__":
    sys.exit(main())
