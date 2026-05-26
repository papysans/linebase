"""Run one truth-set pair through the normal FastAPI job pipeline.

This is intentionally close to the UI flow:
1. build a one-row workbook with HTTP logo/evidence URLs,
2. POST /api/uploads,
3. POST /api/jobs,
4. POST /api/jobs/{id}/start,
5. stream SSE until finished,
6. fetch /api/jobs/{id}/rows and score bbox IoU vs docs/truth_set.

Env:
    LINEBASE_SINGLE_TM=6433801
    LINEBASE_SINGLE_PAIR=1
    LINEBASE_E2E_MODEL=Qwen/Qwen3-VL-30B-A3B-Instruct
    LINEBASE_SINGLE_VERIFY=1
    LINEBASE_SINGLE_TILE_SCAN=0
"""
from __future__ import annotations

import asyncio
import contextlib
import json
import os
import socket
import subprocess
import sys
import time
from pathlib import Path
from typing import Any

import httpx
import openpyxl

REPO = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO / "src"))
sys.stdout.reconfigure(encoding="utf-8", errors="replace")  # type: ignore[union-attr]

OUT_DIR = REPO / "scripts" / "_e2e_out" / "single_truth"
SHEET = "truth_single"


def _pick_free_port() -> int:
    sock = socket.socket()
    sock.bind(("127.0.0.1", 0))
    port = int(sock.getsockname()[1])
    sock.close()
    return port


def _truth_pair(tm: str, pair_idx: int) -> tuple[dict[str, Any], dict[str, Any]]:
    idx = json.loads((REPO / "docs" / "truth_set" / "INDEX.json").read_text(encoding="utf-8"))
    block = next(b for b in idx if b["tm"] == tm)
    pair = next(p for p in block["pairs"] if int(p["i"]) == pair_idx)
    return block, pair


def _write_workbook(path: Path, appno: str, logo_url: str, evidence_url: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET
    ws["B1"] = "appno"
    ws["D1"] = "logo_url"
    ws["K1"] = "evidence_urls"
    ws["B2"] = appno
    ws["D2"] = logo_url
    ws["K2"] = evidence_url
    wb.save(path)


def _wait_for_health(port: int, timeout_s: float = 60.0) -> None:
    deadline = time.time() + timeout_s
    while time.time() < deadline:
        with contextlib.suppress(httpx.HTTPError):
            r = httpx.get(f"http://127.0.0.1:{port}/api/dev/eval-runs", timeout=2.0)
            if r.status_code < 500:
                return
        time.sleep(0.5)
    raise RuntimeError(f"uvicorn at :{port} did not become healthy")


def _wait_for_static(port: int, rel_path: str, timeout_s: float = 20.0) -> None:
    deadline = time.time() + timeout_s
    url = f"http://127.0.0.1:{port}/{rel_path.replace(os.sep, '/')}"
    while time.time() < deadline:
        with contextlib.suppress(httpx.HTTPError):
            r = httpx.get(url, timeout=2.0)
            if r.status_code == 200 and r.content:
                return
        time.sleep(0.3)
    raise RuntimeError(f"asset server did not serve {url}")


async def _drain_events(client: httpx.AsyncClient, job_id: str) -> dict[str, Any]:
    last_job: dict[str, Any] = {}
    t0 = time.time()
    async with client.stream(
        "GET",
        f"/api/jobs/{job_id}/events",
        timeout=httpx.Timeout(900.0, connect=10.0),
    ) as resp:
        resp.raise_for_status()
        buf = ""
        async for chunk in resp.aiter_text():
            buf += chunk.replace("\r\n", "\n")
            while "\n\n" in buf:
                raw, buf = buf.split("\n\n", 1)
                payload = "\n".join(
                    line.removeprefix("data:").strip()
                    for line in raw.splitlines()
                    if line.startswith("data:")
                )
                if not payload:
                    continue
                try:
                    ev = json.loads(payload)
                except json.JSONDecodeError:
                    continue
                et = ev.get("type")
                dt = time.time() - t0
                if et == "progress" and "job" in ev:
                    last_job = ev["job"]
                    print(
                        f"[{dt:6.1f}s] progress {last_job['status']} "
                        f"{last_job['done_rows']}/{last_job['total_rows']} "
                        f"cost=${last_job['cost_usd']:.4f}",
                    )
                elif et == "progress" and "row" in ev:
                    row = ev["row"]
                    print(f"[{dt:6.1f}s] row.start id={row['id']} status={row['status']}")
                elif et in {"row_done", "row_failed"}:
                    row = ev["row"]
                    print(
                        f"[{dt:6.1f}s] {et} status={row['status']} "
                        f"best_conf={row.get('best_confidence')} crop={row.get('best_crop_path')}",
                    )
                elif et == "finished":
                    last_job = ev["job"]
                    print(
                        f"[{dt:6.1f}s] finished {last_job['status']} "
                        f"cost=${last_job['cost_usd']:.4f}",
                    )
                    return last_job
                elif et == "warning":
                    print(f"[{dt:6.1f}s] warning {ev.get('message')}")
    return last_job


def _iou(a: list[int] | tuple[int, int, int, int] | None, b: list[int] | None) -> float:
    if not a or not b:
        return 0.0
    ax1, ay1, ax2, ay2 = (float(v) for v in a)
    bx1, by1, bx2, by2 = (float(v) for v in b)
    ix1, iy1 = max(ax1, bx1), max(ay1, by1)
    ix2, iy2 = min(ax2, bx2), min(ay2, by2)
    iw, ih = max(0.0, ix2 - ix1), max(0.0, iy2 - iy1)
    inter = iw * ih
    area_a = max(0.0, ax2 - ax1) * max(0.0, ay2 - ay1)
    area_b = max(0.0, bx2 - bx1) * max(0.0, by2 - by1)
    union = area_a + area_b - inter
    return inter / union if union > 0 else 0.0


async def _run() -> int:
    tm = os.environ.get("LINEBASE_SINGLE_TM", "6433801").strip()
    pair_idx = int(os.environ.get("LINEBASE_SINGLE_PAIR", "1"))
    model = os.environ.get("LINEBASE_E2E_MODEL", "Qwen/Qwen3-VL-30B-A3B-Instruct").strip()
    verify = os.environ.get("LINEBASE_SINGLE_VERIFY", "1").strip().lower() not in {
        "0",
        "false",
        "off",
        "no",
    }
    tile_scan = os.environ.get("LINEBASE_SINGLE_TILE_SCAN", "0").strip().lower() in {
        "1",
        "true",
        "on",
        "yes",
    }

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    block, pair = _truth_pair(tm, pair_idx)
    logo_rel = str(Path(block["logo"]))
    evidence_rel = str(Path(pair["evidence"]))
    truth = pair["truth_bbox"]

    asset_port = _pick_free_port()
    api_port = _pick_free_port()
    logo_url = f"http://127.0.0.1:{asset_port}/{logo_rel.replace(os.sep, '/')}"
    evidence_url = f"http://127.0.0.1:{asset_port}/{evidence_rel.replace(os.sep, '/')}"
    workbook = OUT_DIR / f"{tm}_pair_{pair_idx:02d}.xlsx"
    _write_workbook(workbook, tm, logo_url, evidence_url)

    asset_log = (OUT_DIR / "asset_http.log").open("wb")
    api_log = (OUT_DIR / "uvicorn.log").open("wb")
    env = dict(os.environ)
    env["PYTHONPATH"] = str(REPO / "src")
    env["PYTHONUNBUFFERED"] = "1"

    print(f"[single-e2e] tm={tm} pair={pair_idx:02d} model={model}")
    print(f"[single-e2e] truth={truth}")
    print(f"[single-e2e] verify={verify} tile_scan={tile_scan}")

    asset_proc = subprocess.Popen(
        [sys.executable, "-u", "-m", "http.server", str(asset_port), "--bind", "127.0.0.1"],
        cwd=str(REPO),
        stdout=asset_log,
        stderr=subprocess.STDOUT,
    )
    api_proc = subprocess.Popen(
        [
            sys.executable,
            "-u",
            "-m",
            "uvicorn",
            "linebase.server:app",
            "--host",
            "127.0.0.1",
            "--port",
            str(api_port),
            "--log-level",
            "warning",
        ],
        cwd=str(REPO),
        env=env,
        stdout=api_log,
        stderr=subprocess.STDOUT,
    )
    try:
        _wait_for_static(asset_port, logo_rel)
        _wait_for_static(asset_port, evidence_rel)
        _wait_for_health(api_port)
        async with httpx.AsyncClient(
            base_url=f"http://127.0.0.1:{api_port}",
            timeout=httpx.Timeout(600.0, connect=10.0),
        ) as client:
            with workbook.open("rb") as fh:
                files = {
                    "file": (
                        workbook.name,
                        fh,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    ),
                }
                upload_resp = await client.post("/api/uploads", files=files, timeout=120.0)
            upload_resp.raise_for_status()
            upload = upload_resp.json()

            job_req = {
                "upload_id": upload["id"],
                "sheet_name": SHEET,
                "appno_column": "B",
                "logo_column": "D",
                "evidence_column": "K",
                "sample_kind": "row_ids",
                "sample_params": {"ids": [2]},
                "threshold": 0.5,
                "model": model,
                "verify_loop": verify,
                "tile_scan": tile_scan,
            }
            job_resp = await client.post("/api/jobs", json=job_req)
            job_resp.raise_for_status()
            job = job_resp.json()
            job_id = job["id"]
            print(f"[single-e2e] job={job_id} rows={job['total_rows']}")

            start_resp = await client.post(f"/api/jobs/{job_id}/start")
            start_resp.raise_for_status()
            await _drain_events(client, job_id)

            rows_resp = await client.get(f"/api/jobs/{job_id}/rows")
            rows_resp.raise_for_status()
            rows = rows_resp.json()
            row = rows[0]
            meta = row.get("match_meta") or {}
            info = next(iter(meta.values())) if meta else {}
            bbox = info.get("bbox") if isinstance(info, dict) else None
            score = _iou(bbox, truth)
            result = {
                "tm": tm,
                "pair": pair_idx,
                "job_id": job_id,
                "model": model,
                "verify_loop": verify,
                "tile_scan": tile_scan,
                "status": row.get("status"),
                "best_crop_path": row.get("best_crop_path"),
                "truth_bbox": truth,
                "pred_bbox": bbox,
                "iou": score,
                "meta": info,
                "row": row,
            }
            out_json = OUT_DIR / "result.json"
            out_json.write_text(json.dumps(result, indent=2, ensure_ascii=False), encoding="utf-8")

            print("[single-e2e] result")
            print(f"  status={row.get('status')} best_crop={row.get('best_crop_path')}")
            print(f"  pred={bbox} truth={truth} iou={score:.3f}")
            if isinstance(info, dict):
                print(
                    "  coord="
                    f"{info.get('bbox_coord_mode')} raw={info.get('raw_bbox')} "
                    f"sent={info.get('sent_size')} source={info.get('source_size')}",
                )
                print(
                    "  verify="
                    f"{info.get('verified')} fit={info.get('fit')} "
                    f"conf={info.get('verify_confidence')} reason={info.get('verify_reason')}",
                )
                print(
                    "  refinements="
                    f"edge={info.get('edge_refined')} tile={info.get('tile_scanned')} "
                    f"pass1_blank={info.get('pass1_blank_reject')}",
                )
            print(f"[single-e2e] wrote {out_json}")
    finally:
        for proc in (api_proc, asset_proc):
            proc.terminate()
        for proc in (api_proc, asset_proc):
            with contextlib.suppress(subprocess.TimeoutExpired):
                proc.wait(timeout=5)
            if proc.poll() is None:
                proc.kill()
        asset_log.close()
        api_log.close()
    return 0


def main() -> int:
    return asyncio.run(_run())


if __name__ == "__main__":
    raise SystemExit(main())
