"""XLSX I/O — read uploaded workbooks, write result workbooks with embedded crops."""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter


@dataclass
class SheetInfo:
    name: str
    rows: int
    columns: int
    header: list[str]
    sample_rows: list[list[str | None]]  # first 5 data rows


def inspect_workbook(path: Path) -> list[SheetInfo]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out: list[SheetInfo] = []
    for name in wb.sheetnames:
        ws = wb[name]
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        rows_iter = ws.iter_rows(min_row=1, max_row=min(6, max_row), values_only=True)
        rows = [[(str(v) if v is not None else None) for v in r] for r in rows_iter]
        header = rows[0] if rows else []
        sample = rows[1:6] if len(rows) > 1 else []
        out.append(SheetInfo(name=name, rows=max_row, columns=max_col, header=[h or "" for h in header], sample_rows=sample))
    wb.close()
    return out


def iter_rows(
    path: Path,
    sheet_name: str,
    appno_col: str,
    logo_col: str,
    evidence_col: str,
    start_row: int = 2,
    end_row: int | None = None,
) -> list[dict[str, str | int | list[str] | None]]:
    """Yield {row_index, appno, logo_url, evidence_urls} for each row in [start_row, end_row]."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    max_row = ws.max_row or start_row
    if end_row is None:
        end_row = max_row
    else:
        end_row = min(end_row, max_row)
    out: list[dict[str, str | int | list[str] | None]] = []
    for r in range(start_row, end_row + 1):
        appno = ws[f"{appno_col}{r}"].value
        logo = ws[f"{logo_col}{r}"].value
        ev = ws[f"{evidence_col}{r}"].value
        ev_list: list[str] = []
        if isinstance(ev, str):
            ev_list = [u.strip() for u in ev.split(",") if u.strip()]
        out.append({
            "row_index": r,
            "appno": str(appno).strip() if appno is not None else None,
            "logo_url": str(logo).strip() if isinstance(logo, str) and logo.startswith("http") else None,
            "evidence_urls": ev_list,
        })
    wb.close()
    return out


def write_result_workbook(
    source_path: Path,
    out_path: Path,
    rows: list[dict],
) -> Path:
    """Open the source read-only; create a new workbook with selected columns + crops embedded.

    For an MVP we don't try to preserve every cell of the 428MB source — we build a fresh result
    workbook with the columns that matter for the reviewer.
    """
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "results"
    ws.append(["row_index", "申请号", "logo_url", "evidence_urls", "status", "crop_path", "confidence", "notes"])
    for r in rows:
        ws.append([
            r["row_index"], r.get("appno"), r.get("logo_url"),
            ",".join(r.get("evidence_urls", [])),
            r.get("status", "pending"),
            r.get("best_crop_path") or "",
            r.get("confidence") or "",
            r.get("notes") or "",
        ])
    # embed the best crop image into column I for each row
    image_col = "I"
    ws.cell(row=1, column=9, value="预览")
    for idx, r in enumerate(rows, start=2):
        crop = r.get("best_crop_path")
        if crop and Path(crop).exists():
            img = XLImage(crop)
            img.width = min(200, img.width)
            img.height = min(200, img.height)
            ws.row_dimensions[idx].height = 120
            ws.column_dimensions[image_col].width = 28
            ws.add_image(img, f"{image_col}{idx}")
    wb.save(out_path)
    return out_path
