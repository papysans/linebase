"""FastAPI app: REST API + SSE + static SPA mount."""
from __future__ import annotations

import io
import json
import time
import zipfile
from pathlib import Path

from fastapi import FastAPI, HTTPException, UploadFile, File, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse, Response, StreamingResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import load_workbook
from pydantic import BaseModel
from sse_starlette.sse import EventSourceResponse

from linebase import store
from linebase.config import Settings
from linebase.fetch import fetch
from linebase.io_excel import inspect_workbook, iter_rows, write_result_workbook
from linebase.models_catalog import MODEL_WHITELIST, is_model_routable, to_dict as model_to_dict
from linebase.pipeline_runner import (
    _active_tasks,
    _job_to_dict,
    _row_to_dict,
    publish,
    start_job,
    stream_events,
)

REPO = Path(__file__).resolve().parents[2]
DATA_DIR = store.DATA_DIR
UPLOADS_DIR = DATA_DIR / "uploads"
RUNS_DIR = DATA_DIR / "runs"
STATIC_DIR = Path(__file__).resolve().parent / "static"


app = FastAPI(title="linebase", version="0.1.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://127.0.0.1:5173"],
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
def _startup() -> None:
    store.init_schema()
    UPLOADS_DIR.mkdir(parents=True, exist_ok=True)
    RUNS_DIR.mkdir(parents=True, exist_ok=True)


# --- Models catalog --------------------------------------------------------


@app.get("/api/models")
def list_models() -> dict:
    """Return the curated model whitelist + the system default + a custom-allowed flag.

    The frontend renders this as a dropdown with a "custom" escape hatch. The
    server validates the chosen model on POST /api/jobs via `is_model_routable`.
    """
    settings = Settings.from_env()
    return {
        "whitelist": [model_to_dict(opt) for opt in MODEL_WHITELIST],
        "default": settings.model,
        "allow_custom": True,
    }


# --- Upload ----------------------------------------------------------------

class UploadResponse(BaseModel):
    id: str
    filename: str
    size: int
    sheets: list[dict]


@app.post("/api/uploads", response_model=UploadResponse)
async def post_upload(file: UploadFile = File(...)) -> UploadResponse:
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(400, "expected .xlsx file")
    UPLOADS_DIR.mkdir(parents=True, exist_ok=True)
    body = await file.read()
    uid = store.new_id()
    target = UPLOADS_DIR / f"{uid}_{file.filename}"
    target.write_bytes(body)
    upload = store.insert_upload(filename=file.filename, size=len(body), path=str(target))
    sheets = [s.__dict__ for s in inspect_workbook(target)]
    store.set_upload_sheets(upload.id, json.dumps(sheets, ensure_ascii=False))
    return UploadResponse(id=upload.id, filename=upload.filename, size=upload.size, sheets=sheets)


@app.get("/api/uploads/{upload_id}", response_model=UploadResponse)
def get_upload(upload_id: str) -> UploadResponse:
    u = store.get_upload(upload_id)
    if not u:
        raise HTTPException(404, "upload not found")
    sheets = json.loads(u.sheets_json or "[]")
    return UploadResponse(id=u.id, filename=u.filename, size=u.size, sheets=sheets)


# --- Jobs ------------------------------------------------------------------

class CreateJobRequest(BaseModel):
    upload_id: str
    sheet_name: str
    appno_column: str
    logo_column: str
    evidence_column: str
    sample_kind: str  # "first_n" | "range" | "row_ids"
    sample_params: dict
    threshold: float = 0.5
    # Per-job model override. None → fall back to Settings.from_env().model.
    # Validated in create_job(): must be in MODEL_WHITELIST or routable via
    # _PROVIDER_PREFIXES — otherwise 400.
    model: str | None = None
    # Default-on: the pipeline runs the verify-loop (extra LLM call per
    # evidence to confirm the crop) — slower + ~2x cost but catches the
    # "brand-recognition shortcut" failure (job 2a2e801827dc457b row 79 picked
    # the Heat fireball when the TM was a basketball-player silhouette).
    # Per the review-loop policy in memory/feedback_review_loop.md: prod batches
    # default ON; cost overhead is worth catching wrong-identity false-positives.
    verify_loop: bool = True
    # Iter 6.3 — opt-in 3x3 tile-scan fallback for small logos buried in busy
    # photos. When True, the pipeline tiles each evidence whose longest side
    # > 1500 px into a 3x3 grid and re-tries the match per tile after the
    # primary path fails / rejects verification. Costs 9 extra LLM calls per
    # affected evidence. Default OFF for backward compat.
    tile_scan: bool = False


def _resolve_rows(upload: store.Upload, req: CreateJobRequest) -> list[dict]:
    path = Path(upload.path)
    if req.sample_kind == "first_n":
        n = int(req.sample_params.get("n", 10))
        # rows below header. Header detection: assume row 1 = header, data starts at row 2.
        # But some sheets (图形商标tro) have row 2 also as a Chinese header — we skip rows 2-3 if hidden.
        rows = iter_rows(path, req.sheet_name, req.appno_column, req.logo_column, req.evidence_column,
                         start_row=2, end_row=2 + n - 1 + 2)
        usable = [r for r in rows if r["logo_url"] or r["evidence_urls"]]
        if not usable and _row_has_selected_inputs(path, req, 1):
            return iter_rows(path, req.sheet_name, req.appno_column, req.logo_column, req.evidence_column,
                             start_row=1, end_row=n)
        return rows
    if req.sample_kind == "range":
        start = int(req.sample_params["start"])
        end = int(req.sample_params["end"])
        return iter_rows(path, req.sheet_name, req.appno_column, req.logo_column, req.evidence_column,
                         start_row=start, end_row=end)
    if req.sample_kind == "row_ids":
        ids: list[int] = req.sample_params["ids"]
        start = min(ids); end = max(ids)
        all_rows = iter_rows(path, req.sheet_name, req.appno_column, req.logo_column, req.evidence_column,
                             start_row=start, end_row=end)
        return [r for r in all_rows if r["row_index"] in ids]
    raise HTTPException(400, f"unknown sample_kind: {req.sample_kind}")


def _row_has_selected_inputs(path: Path, req: CreateJobRequest, row_idx: int) -> bool:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[req.sheet_name]
        logo = ws[f"{req.logo_column}{row_idx}"].value
        ev = ws[f"{req.evidence_column}{row_idx}"].value
    finally:
        wb.close()
    return (
        isinstance(logo, str)
        and logo.strip().startswith("http")
        and isinstance(ev, str)
        and "http" in ev
    )


@app.post("/api/jobs")
def create_job(req: CreateJobRequest) -> dict:
    upload = store.get_upload(req.upload_id)
    if not upload:
        raise HTTPException(404, "upload not found")
    # Reject unroutable model ids early so users see the failure at job-creation
    # time, not as a per-row LLM error 5 minutes into the run.
    if req.model is not None and req.model.strip():
        candidate = req.model.strip()
        if not is_model_routable(candidate):
            raise HTTPException(
                400,
                f"model {candidate!r} is not in the whitelist and does not match any routable provider prefix",
            )
        model_value: str | None = candidate
    else:
        model_value = None
    rows_data = _resolve_rows(upload, req)
    # filter out blatantly empty rows (no logo URL AND no evidence)
    rows_data = [r for r in rows_data if r["logo_url"] or r["evidence_urls"]]
    if not rows_data:
        raise HTTPException(
            400,
            "no usable rows found: check header row and selected app/logo/evidence columns",
        )
    job = store.insert_job(
        upload_id=req.upload_id, sheet_name=req.sheet_name,
        logo_column=req.logo_column, evidence_column=req.evidence_column,
        appno_column=req.appno_column, threshold=req.threshold,
        sample_kind=req.sample_kind, sample_params=req.sample_params,
        model=model_value,
        verify_loop=1 if req.verify_loop else 0,
        tile_scan=1 if req.tile_scan else 0,
    )
    for r in rows_data:
        store.insert_job_row(
            job_id=job.id, row_index=int(r["row_index"]),
            appno=r.get("appno"), logo_url=r.get("logo_url"),
            evidence_urls=r["evidence_urls"] or [],
        )
    store.update_job(job.id, total_rows=len(rows_data))
    return _job_to_dict(store.get_job(job.id))  # type: ignore[arg-type]


@app.post("/api/jobs/{job_id}/cancel")
async def cancel(job_id: str) -> dict:
    """Cancel an in-flight job.

    Effects:
      1. If `_active_tasks[job_id]` holds a not-done asyncio.Task, call
         `.cancel()` on it. The task's awaiters (run_in_executor) cope with
         CancelledError by propagating up — `_run_job` then exits its loop.
      2. Mark the job row `status="cancelled"` in SQLite.
      3. Flip any row still in `pending` or `running` to `failed` with
         `notes="cancelled by user"` — distinct from a real `failed` because
         the notes field tells the reviewer it was a deliberate stop.
      4. Publish a `cancelled` SSE event so the RunPage stops the spinner.

    Idempotent: re-cancelling an already-cancelled job is a no-op that still
    returns the current job snapshot.
    """
    job = store.get_job(job_id)
    if not job:
        raise HTTPException(404)

    # 1. Kill the asyncio task if it's running.
    task = _active_tasks.get(job_id)
    if task is not None and not task.done():
        task.cancel()

    # 2 + 3. Persist cancellation state.
    store.update_job(job_id, status="cancelled")
    for r in store.list_job_rows(job_id):
        if r.status in {"pending", "running"}:
            store.update_job_row(r.id, status="failed", notes="cancelled by user")

    # 4. Fire-and-forget SSE notification. We re-fetch the job so the event
    # carries the post-cancel state.
    refreshed = store.get_job(job_id)
    if refreshed is not None:
        await publish(job_id, {"type": "progress", "job": _job_to_dict(refreshed)})
        await publish(
            job_id,
            {"type": "warning", "message": "任务已被用户取消"},
        )

    return _job_to_dict(refreshed)  # type: ignore[arg-type]


@app.post("/api/jobs/{job_id}/start")
async def start(job_id: str) -> dict:
    # NB: async def is required — start_job calls asyncio.get_event_loop().
    # When this route was sync, FastAPI dispatched it on the anyio thread pool
    # where get_event_loop() raises "no current event loop in thread 'AnyIO
    # worker thread'" on Python 3.11+. See e2e_real_xlsx.py for the repro.
    job = store.get_job(job_id)
    if not job:
        raise HTTPException(404)
    upload = store.get_upload(job.upload_id)
    assert upload is not None
    settings = Settings.from_env()
    start_job(job_id, Path(upload.path), settings)
    return _job_to_dict(store.get_job(job_id))  # type: ignore[arg-type]


@app.get("/api/jobs")
def list_jobs_view(limit: int = 50) -> list[dict]:
    """Recent jobs, newest first. Backs the frontend empty-state on
    /run, /review, /download when the user lands without a jobId in URL.

    Clamped to [1, 200] to keep the response small (and because the picker
    only shows the top ~5 anyway).
    """
    limit = max(1, min(200, int(limit)))
    return [_job_to_dict(j) for j in store.list_jobs(limit=limit)]


@app.get("/api/jobs/{job_id}")
def get_job_view(job_id: str) -> dict:
    job = store.get_job(job_id)
    if not job:
        raise HTTPException(404)
    return _job_to_dict(job)


@app.get("/api/jobs/{job_id}/rows")
def list_rows(job_id: str, status: str | None = None) -> list[dict]:
    rows = store.list_job_rows(job_id, status=status)
    return [_row_to_dict(r) for r in rows]


class SetStatusRequest(BaseModel):
    human_status: str  # "ok" | "bad" | "needs_review"
    notes: str | None = None


@app.post("/api/jobs/{job_id}/rows/{row_id}/status")
def set_row_status(job_id: str, row_id: int, req: SetStatusRequest) -> dict:
    row = store.get_job_row(row_id)
    if not row or row.job_id != job_id:
        raise HTTPException(404)
    store.update_job_row(row_id, human_status=req.human_status, notes=req.notes)
    return _row_to_dict(store.get_job_row(row_id))  # type: ignore[arg-type]


class RerunRequest(BaseModel):
    row_ids: list[int] | None = None


@app.post("/api/jobs/{job_id}/rerun")
async def rerun(job_id: str, req: RerunRequest) -> dict:
    job = store.get_job(job_id)
    if not job:
        raise HTTPException(404)
    # reset selected rows to pending; if no list, reset all rows that aren't human-OK
    rows = store.list_job_rows(job_id)
    targets = [r for r in rows if (req.row_ids is None and r.human_status != "ok") or (req.row_ids and r.id in req.row_ids)]
    for r in targets:
        store.update_job_row(r.id, status="pending", best_crop_path=None, all_crops_json="{}", match_meta_json="{}")
    store.update_job(job_id, status="pending")
    upload = store.get_upload(job.upload_id)
    assert upload is not None
    settings = Settings.from_env()
    start_job(job_id, Path(upload.path), settings)
    return _job_to_dict(store.get_job(job_id))  # type: ignore[arg-type]


class RowRerunRequest(BaseModel):
    """One-shot rerun knobs scoped to a single row. Both fields are opt-in:
    omit them to reuse the job's existing model + verify_loop settings.
    """
    verify: bool = False  # force verify-loop on for this rerun (cost ~2x)
    model: str | None = None  # transient model override for this rerun


@app.post("/api/jobs/{job_id}/rows/{row_id}/rerun")
async def rerun_one_row(job_id: str, row_id: int, req: RowRerunRequest) -> dict:
    """Re-run a single row with optional verify+model overrides.

    Semantics:
      1. Reset the target row to `pending` and clear its crop / meta / human
         status so the next run starts clean.
      2. If the caller asked for verify or a model override, patch those onto
         the job row (these stick — there's no UI to unset them later; the
         simpler "no transient state" path matches the autonomous-loop spirit
         of this project).
      3. Kick off `start_job`. `_run_job` skips ok/bad/needs_review/failed rows,
         so only the just-reset row will be processed.

    Why we don't snapshot-and-restore the job: per-row rerun is an interactive
    debugging tool. The user is iterating on a row that went wrong; if they
    later want the original model back, they re-pick it in the row's rerun
    dialog. Worth one column of muscle-memory for zero state-machine complexity.
    """
    row = store.get_job_row(row_id)
    if not row or row.job_id != job_id:
        raise HTTPException(404)
    job = store.get_job(job_id)
    if not job:
        raise HTTPException(404)

    # Patch job-level overrides first so _process_row sees them when it runs.
    job_updates: dict = {}
    if req.verify:
        job_updates["verify_loop"] = 1
    if req.model and req.model.strip():
        candidate = req.model.strip()
        if not is_model_routable(candidate):
            raise HTTPException(
                400,
                f"model {candidate!r} is not in the whitelist and does not match any routable provider prefix",
            )
        job_updates["model"] = candidate
    if job_updates:
        store.update_job(job_id, **job_updates)

    # Reset the row so the runner picks it up. clear best_crop / crops / meta
    # so the previous (wrong) result doesn't bleed into the UI mid-rerun.
    store.update_job_row(
        row_id,
        status="pending",
        best_crop_path=None,
        all_crops_json="{}",
        match_meta_json="{}",
        human_status=None,
        notes=None,
    )
    # Job-level status flip so the SSE stream and the job summary correctly
    # report "running" while the one row is in-flight.
    store.update_job(job_id, status="pending")

    upload = store.get_upload(job.upload_id)
    assert upload is not None
    settings = Settings.from_env()
    start_job(job_id, Path(upload.path), settings)
    return _job_to_dict(store.get_job(job_id))  # type: ignore[arg-type]


@app.get("/api/jobs/{job_id}/events")
async def events(job_id: str, request: Request) -> EventSourceResponse:
    async def gen():
        async for ev in stream_events(job_id):
            if await request.is_disconnected():
                return
            yield {"data": json.dumps(ev, ensure_ascii=False)}
    return EventSourceResponse(gen())


@app.get("/api/jobs/{job_id}/xlsx")
def download_xlsx(job_id: str) -> FileResponse:
    job = store.get_job(job_id)
    if not job:
        raise HTTPException(404)
    upload = store.get_upload(job.upload_id)
    assert upload is not None
    rows = store.list_job_rows(job_id)
    rows_data = []
    for r in rows:
        meta = json.loads(r.match_meta_json or "{}")
        # find confidence for best_crop_path
        conf = 0.0
        for m in meta.values():
            if isinstance(m, dict) and m.get("confidence", 0) > conf:
                conf = m.get("confidence", 0)
        rows_data.append({
            "row_index": r.row_index, "appno": r.appno, "logo_url": r.logo_url,
            "evidence_urls": json.loads(r.evidence_urls_json),
            "status": r.human_status or r.status,
            "best_crop_path": r.best_crop_path, "confidence": conf, "notes": r.notes or "",
        })
    out = RUNS_DIR / job_id / f"result_{int(time.time())}.xlsx"
    write_result_workbook(Path(upload.path), out, rows_data)
    return FileResponse(out, filename=out.name, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.get("/api/jobs/{job_id}/images.zip")
def download_images(job_id: str, status: str | None = None) -> StreamingResponse:
    rows = store.list_job_rows(job_id)
    if status:
        # Normalise: frontend may send "OK" / "BAD" / "NEEDS_REVIEW"; the DB
        # stores lowercase ("ok" / "bad" / "needs_review").
        wanted = status.strip().lower()
        rows = [r for r in rows if (r.human_status or r.status or "").lower() == wanted]
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        for r in rows:
            crops = json.loads(r.all_crops_json or "{}")
            idx = 0
            for url, path in crops.items():
                if not path or not Path(path).exists():
                    continue
                idx += 1
                ext = Path(path).suffix or ".png"
                z.write(path, f"{r.appno or r.id}_{idx}{ext}")
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/zip",
                             headers={"Content-Disposition": f"attachment; filename=images_{job_id}.zip"})


@app.get("/api/jobs/{job_id}/file")
def get_job_file(job_id: str, p: str) -> Response:
    # safe-serve: only allow paths within RUNS_DIR/<job_id>/
    target = Path(p).resolve()
    base = (RUNS_DIR / job_id).resolve()
    if not str(target).startswith(str(base)):
        raise HTTPException(403)
    if not target.exists():
        raise HTTPException(404)
    return FileResponse(target)


@app.get("/api/img")
def get_image(u: str) -> Response:
    """Proxy + cache: fetch the URL and return its bytes, used by frontend to display logos."""
    path = fetch(u)
    return FileResponse(path)


def _enrich_eval_metrics(raw: dict) -> dict:
    """Project raw eval-run JSON into the metric shape the frontend reads.

    The eval harness writes `mean_ssim`, `mean_phash`, `pass_rate_ssim_50`,
    `cost_usd_estimate`, and a `pairs` list with per-sample `selection_correct`
    / `iou_vs_redbox` fields. The frontend reads `selection_accuracy`,
    `mean_iou`, `mean_ssim`, `cost_usd`, `pass_rate`, `n_samples`. Compute the
    missing aggregates here so the leaderboard and the metric cards aren't
    full of `—` placeholders.
    """
    out = dict(raw or {})
    pairs = out.get("pairs") or []
    # selection_accuracy = (# pairs with selection_correct=True) / (# pairs
    # where selection_correct is not None). `None` means the pair had no
    # ground-truth red-box, so it shouldn't count for/against accuracy.
    sels = [p.get("selection_correct") for p in pairs]
    judged = [s for s in sels if s is not None]
    if judged:
        out.setdefault("selection_accuracy", sum(1 for s in judged if s) / len(judged))
    # mean_iou over pairs where iou_vs_redbox is a number
    ious = [p.get("iou_vs_redbox") for p in pairs if isinstance(p.get("iou_vs_redbox"), (int, float))]
    if ious:
        out.setdefault("mean_iou", sum(ious) / len(ious))
    # Frontend-friendly aliases
    if "cost_usd" not in out and "cost_usd_estimate" in out:
        out["cost_usd"] = out["cost_usd_estimate"]
    if "pass_rate" not in out and "pass_rate_ssim_50" in out:
        out["pass_rate"] = out["pass_rate_ssim_50"]
    if "n_samples" not in out and "samples" in out:
        out["n_samples"] = out["samples"]
    return out


@app.get("/api/dev/eval-runs")
def eval_runs() -> list[dict]:
    runs = store.list_eval_runs()
    return [{"id": r["id"], "prompt_version": r["prompt_version"], "model": r["model"],
             "metrics": _enrich_eval_metrics(json.loads(r["metrics_json"])),
             "created_at": r["created_at"]} for r in runs]


# --- Static SPA ------------------------------------------------------------

if STATIC_DIR.exists():
    app.mount("/assets", StaticFiles(directory=str(STATIC_DIR / "assets")), name="assets")

    @app.get("/{rest:path}")
    def spa(rest: str) -> Response:
        # Guard: never serve the SPA index for /api/* paths. Without this,
        # malformed API URLs (e.g. /api/jobs//xlsx when jobId is empty, or any
        # typo'd endpoint) silently return index.html. The browser then saves
        # an HTML file under the .xlsx URL extension, and the user complains
        # the "downloaded file isn't xlsx". Explicit 404 makes the failure
        # mode loud at the network layer.
        if rest.startswith("api/") or rest == "api":
            raise HTTPException(404)
        index = STATIC_DIR / "index.html"
        if not index.exists():
            raise HTTPException(404)
        return FileResponse(index)
